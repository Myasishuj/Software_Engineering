from flask import Flask, request, jsonify, send_file, make_response
from flask_cors import CORS
from flask_jwt_extended import create_access_token, jwt_required, JWTManager, get_jwt_identity
import pandas as pd
import io
import json
from datetime import datetime, timedelta, timezone
import logging
from functools import wraps 
import base64 # For base64 encoding
from barcode import UPCA, Code128 # Import UPCA, keep Code128 as fallback/reference
from barcode.writer import SVGWriter # For SVG output
import hashlib # For creating a stable 11-digit number from pid

# For MongoDB
from pymongo import MongoClient
from bson.objectid import ObjectId 
from werkzeug.security import generate_password_hash, check_password_hash

# For Email Notifications
import smtplib
from email.message import EmailMessage
from socket import gaierror # Import for network-related errors

# Import configurations from config.py
from config import (
    JWT_SECRET_KEY, JWT_ACCESS_TOKEN_EXPIRES,
    MONGO_URI, DB_NAME, TEMPLATES_COLLECTION, USERS_COLLECTION, SUBMISSIONS_COLLECTION,
    SMTP_SERVER, SMTP_PORT, SENDER_EMAIL, SENDER_PASSWORD,
    ENABLE_ALL_USERS_DAILY_DOWNLOAD, ENABLE_ALL_USERS_GLOBAL_SEARCH # Import new toggles
)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Define the allowed origin for your frontend.
ALLOWED_ORIGIN = "http://localhost:5173"

# Initialize Flask-CORS with basic settings.
CORS(app) 

# --- ULTIMATE BRUTE FORCE CORS HANDLING FOR OPTIONS REQUESTS ---
@app.before_request
def handle_options_preflight():
    if request.method == 'OPTIONS':
        print(f"DEBUG: Intercepted OPTIONS request for path: {request.path}")
        response = make_response()
        response.headers.add('Access-Control-Allow-Origin', ALLOWED_ORIGIN)
        response.headers.add('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type, Authorization, X-Requested-With')
        response.headers.add('Access-Control-Allow-Credentials', 'true')
        response.headers.add('Access-Control-Max-Age', '86400') 
        return response, 200

# --- GENERIC OPTIONS ROUTE AS A FALLBACK (Added for robustness) ---
@app.route('/<path:path>', methods=['OPTIONS'])
def catch_all_options(path):
    print(f"DEBUG: Hit generic OPTIONS route for path: /{path}")
    response = make_response()
    response.headers.add('Access-Control-Allow-Origin', ALLOWED_ORIGIN)
    response.headers.add('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type, Authorization, X-Requested-With')
    response.headers.add('Access-Control-Allow-Credentials', 'true')
    response.headers.add('Access-Control-Max-Age', '86400')
    return response, 200


# --- JWT Configuration ---
app.config["JWT_SECRET_KEY"] = JWT_SECRET_KEY
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = JWT_ACCESS_TOKEN_EXPIRES 
jwt = JWTManager(app)

# --- MongoDB Connection ---
client = MongoClient(MONGO_URI)
db = client[DB_NAME] 
templates_collection = db[TEMPLATES_COLLECTION]
users_collection = db[USERS_COLLECTION] 
submissions_collection = db[SUBMISSIONS_COLLECTION]

# --- Helper function for sending emails ---
def send_email_notification(recipient_email, subject, body):
    msg = EmailMessage()
    msg.set_content(body)
    msg['Subject'] = subject
    msg['From'] = SENDER_EMAIL
    msg['To'] = recipient_email

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
            smtp.starttls() 
            smtp.login(SENDER_EMAIL, SENDER_PASSWORD)
            smtp.send_message(msg)
        logger.info(f"Email sent successfully to {recipient_email}")
        return True
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"Failed to send email to {recipient_email}: Authentication error. Check SENDER_EMAIL and SENDER_PASSWORD. Details: {e}")
        return False
    except smtplib.SMTPConnectError as e:
        logger.error(f"Failed to send email to {recipient_email}: Connection error. Check SMTP_SERVER and SMTP_PORT. Details: {e}")
        return False
    except gaierror as e: # Catch DNS resolution errors
        logger.error(f"Failed to send email to {recipient_email}: DNS resolution error. Check SMTP_SERVER. Details: {e}")
        return False
    except Exception as e:
        logger.error(f"Failed to send email to {recipient_email}: An unexpected error occurred: {e}")
        return False

# --- JWT Callbacks (for user identity) ---
@jwt.user_identity_loader
def user_identity_callback(user):
    if isinstance(user, dict) and 'username' in user:
        return user['username']
    return None

@jwt.user_lookup_loader
def user_lookup_callback(_jwt_header, jwt_data):
    identity = jwt_data["sub"]
    return users_collection.find_one({"username": identity})

# --- Custom Role-Based Access Control Decorator ---
def roles_required(roles):
    def decorator(fn):
        @wraps(fn)
        @jwt_required()
        def wrapper(*args, **kwargs):
            current_user_identity = get_jwt_identity()
            user = users_collection.find_one({"username": current_user_identity})
            
            if not user:
                return jsonify({"msg": "User not found."}), 404
            
            if user.get('role') not in roles:
                return jsonify({"msg": "Forbidden: Insufficient permissions."}), 403
            
            return fn(*args, **kwargs)
        return wrapper
    return decorator

# --- Admin role required decorator (now using roles_required) ---
def admin_required(f):
    @wraps(f)
    @roles_required(['admin'])
    def decorated_function(*args, **kwargs):
        return f(*args, **kwargs)
    return decorated_function

# --- Add some default templates and default users to MongoDB if collections are empty ---
def initialize_db(test_mode=False):
    try:
        # Template initialization
        if templates_collection.count_documents({}) == 0:
            logger.info("Initializing default templates...")
            default_templates = [
                {
                    'id': 'sales_report',
                    'name': 'Sales Report',
                    'description': 'Basic sales data template',
                    'columns': ['Product', 'Quantity', 'Price', 'Total'],
                    'sample_data': [
                        {'Product': 'Laptop', 'Quantity': 5, 'Price': 999.99, 'Total': 4999.95},
                        {'Product': 'Mouse', 'Quantity': 20, 'Price': 25.50, 'Total': 510.00}
                    ]
                },
                {
                    'id': 'inventory',
                    'name': 'Inventory List',
                    'description': 'Product inventory tracking',
                    'columns': ['Item', 'SKU', 'Stock', 'Location', 'expiry'],
                    'sample_data': [
                        {'Item': 'Widget A', 'SKU': 'WA001', 'Stock': 155, 'Location': 'Warehouse A', 'expiry': '15-07-2025'},
                        {'Item': 'Widget B', 'SKU': 'WB001', 'Stock': 75, 'Location': 'Warehouse B', 'expiry': '01-01-2026'},
                        {'Item': 'Expired Item C', 'SKU': 'EXC001', 'Stock': 10, 'Location': 'Shelf 3', 'expiry': '10-06-2025'}
                    ]
                }
            ]
            result = templates_collection.insert_many(default_templates)
            if not result.acknowledged:
                raise Exception("Failed to insert default templates.")
            logger.info("Default templates initialized.")
        else:
            logger.info("Templates collection is not empty, skipping initialization.")

        # User initialization
        if users_collection.count_documents({}) == 0:
            logger.info("Initializing default users...")
            default_users = [
                {'username': 'admin', 'password': generate_password_hash('admin'), 'role': 'admin', 'email': 'admin@example.com'},
                {'username': 'user', 'password': generate_password_hash('user'), 'role': 'user', 'email': 'user@example.com'}
            ]
            result = users_collection.insert_many(default_users)
            if not result.acknowledged:
                raise Exception("Failed to insert default users.")
            logger.info("Default users initialized (admin, user).")
        else:
            logger.info("Users collection is not empty, skipping basic user initialization.")

        # Ensure tester1 exists
        tester1_user = users_collection.find_one({"username": "tester1"})
        if not tester1_user:
            result = users_collection.insert_one({
                'username': 'tester1',
                'password': generate_password_hash('tester1'),
                'role': 'tester',
                'email': 'chenhayik@gmail.com'
            })
            if not result.acknowledged:
                raise Exception("Failed to insert tester1 user.")
            logger.info("Default tester1 user created with email chenhayik@gmail.com.")
        else:
            if tester1_user.get('email') != 'chenhayik@gmail.com':
                users_collection.update_one(
                    {"_id": tester1_user['_id']},
                    {"$set": {"email": 'chenhayik@gmail.com'}}
                )
                logger.info("Updated email for existing tester1 user to chenhayik@gmail.com.")

        # Patch missing emails
        for user_doc in users_collection.find({"email": {"$exists": False}}):
            result = users_collection.update_one(
                {"_id": user_doc['_id']},
                {"$set": {"email": f"{user_doc['username']}@example.com"}}
            )
            if result.modified_count != 1:
                raise Exception(f"Failed to update email for user {user_doc['username']}.")
            logger.info(f"Added default email for existing user: {user_doc['username']}")

        # Optional cleanup (for test-only runs)
        if test_mode:
            logger.info("Test mode active: cleaning up initialized data...")
            templates_collection.delete_many({})
            users_collection.delete_many({"username": {"$in": ["admin", "user", "tester1"]}})
            logger.info("Test data removed successfully.")

    except Exception as e:
        logger.error(f"Error during database initialization: {e}")
        raise  # Re-raise the error so it can be caught by the application or test framework


# Call initialization on app startup
with app.app_context():
    initialize_db()


@app.route('/', methods=['GET'])
def home():
    """Health check endpoint"""
    return jsonify({
        'message': 'Excel Creator API is running',
        'status': 'healthy',
        'timestamp': datetime.now().isoformat()
    })

# --- Authentication Routes ---
@app.route('/login', methods=['POST', 'OPTIONS'])
def login():
    if request.method == 'OPTIONS':
        return '', 200 

    username = request.json.get('username', None)
    password = request.json.get('password', None)

    if not username or not password:
        return jsonify({"msg": "Missing username or password"}), 400

    user = users_collection.find_one({"username": username})

    if user and check_password_hash(user['password'], password):
        access_token = create_access_token(identity=user)
        return jsonify(access_token=access_token, username=user['username'], role=user['role']), 200
    else:
        return jsonify({"msg": "Bad username or password"}), 401

@app.route('/register', methods=['POST'])
def register():
    username = request.json.get('username', None)
    password = request.json.get('password', None)
    role = request.json.get('role', 'user') 

    if not username or not password:
        return jsonify({"msg": "Missing username or password"}), 400

    if users_collection.find_one({"username": username}):
        return jsonify({"msg": "Username already exists"}), 409

    hashed_password = generate_password_hash(password)
    email = request.json.get('email', f"{username}@example.com") 
    users_collection.insert_one({"username": username, "password": hashed_password, "role": role, "email": email})

    return jsonify({"msg": "User registered successfully"}), 201

# --- Excel Creation & Template Routes ---
@app.route('/create-excel', methods=['POST'])
@roles_required(['user', 'admin', 'tester']) 
def create_excel():
    """
    Create an Excel file from JSON data and return it as a downloadable file
    This is for direct creation, not tied to approval workflow
    """
    try:
        current_user_identity = get_jwt_identity() 
        logger.info(f"User {current_user_identity} requesting direct Excel creation.")

        json_data = request.get_json()
        
        if not json_data:
            return jsonify({'error': 'No JSON data provided'}), 400
        
        filename = json_data.get('filename', 'data')
        sheet_name = json_data.get('sheet_name', 'Sheet1')
        data = json_data.get('data', [])
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        if not isinstance(data, list):
            return jsonify({'error': 'Data must be a list of objects'}), 400
        
        for i, item in enumerate(data):
            if not isinstance(item, dict):
                return jsonify({'error': f'Item at index {i} is not a valid object'}), 400
        
        df = pd.DataFrame(data)
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
        
        excel_buffer.seek(0)
        safe_filename = f"{filename}.xlsx"
        
        logger.info(f"Created Excel file: {safe_filename} for user {current_user_identity} with {len(data)} rows")
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=safe_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error creating Excel file for user {current_user_identity}: {str(e)}")
        return jsonify({'error': f'Failed to create Excel file: {str(e)}'}), 500

@app.route('/validate-data', methods=['POST'])
@roles_required(['user', 'admin', 'tester']) 
def validate_data():
    """
    Validate JSON data structure without creating the Excel file
    """
    try:
        json_data = request.get_json()
        
        if not json_data:
            return jsonify({'error': 'No JSON data provided'}), 400
        
        data = json_data.get('data', [])
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        if not isinstance(data, list):
            return jsonify({'error': 'Data must be a list of objects'}), 400
        
        columns = set()
        for i, item in enumerate(data):
            if not isinstance(item, dict):
                return jsonify({'error': f'Item at index {i} is not a valid object'}), 400
            columns.update(item.keys())
        
        df = pd.DataFrame(data)
        
        return jsonify({
            'valid': True,
            'row_count': len(data),
            'columns': list(columns),
            'preview': df.head().to_dict('records') if len(data) > 0 else []
        })
        
    except Exception as e:
        logger.error(f"Error validating data: {str(e)}")
        return jsonify({'error': f'Validation failed: {str(e)}'}), 500

@app.route('/templates', methods=['GET'])
@roles_required(['user', 'admin', 'tester']) 
def get_templates():
    """
    Get available Excel templates from MongoDB
    """
    try:
        templates_cursor = templates_collection.find({}, {'_id': 0}) 
        templates = list(templates_cursor)
        return jsonify({'templates': templates})
    except Exception as e:
        logger.error(f"Error fetching templates from MongoDB: {str(e)}")
        return jsonify({'error': 'Failed to retrieve templates'}), 500

@app.route('/create-from-template/<template_id>', methods=['POST'])
@roles_required(['user', 'admin', 'tester']) 
def create_from_template(template_id):
    """
    Create an Excel file from a predefined template stored in MongoDB
    """
    try:
        current_user_identity = get_jwt_identity()
        template = templates_collection.find_one({'id': template_id}, {'_id': 0})

        if not template:
            return jsonify({'error': 'Template not found'}), 404

        filename = template.get('name', 'template_data').lower().replace(' ', '_')
        sheet_name = template.get('sheet_name', template.get('name', 'Sheet1'))
        data = template.get('sample_data', []) 

        if not data:
            return jsonify({'error': 'Template has no sample data'}), 400

        df = pd.DataFrame(data)
        excel_buffer = io.BytesIO()

        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            from openpyxl.styles import Font, PatternFill, Border, Side
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

        excel_buffer.seek(0)
        safe_filename = f"{filename}.xlsx"

        logger.info(f"Created Excel file from template '{template_id}' for user {current_user_identity}: {safe_filename} with {len(data)} rows")

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=safe_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Error creating Excel file from template {template_id} for user {current_user_identity}: {str(e)}")
        return jsonify({'error': f'Failed to create Excel file from template: {str(e)}'}), 500

# --- Daily Data Submission & Approval Workflow Endpoints ---

@app.route('/dashboard/submit-data', methods=['POST'])
@roles_required(['user', 'tester']) 
def submit_daily_data():
    """
    User endpoint to submit data for admin approval.
    Data is stored in submissions_collection with status "pending".
    """
    logger.info(f"Received request to submit daily data.")
    current_user_identity = get_jwt_identity() 
    user = users_collection.find_one({"username": current_user_identity})

    if not user:
        logger.warning(f"User '{current_user_identity}' not found during data submission attempt.")
        return jsonify({"msg": "User not found."}), 404
    
    records = request.get_json()
    
    if not records or not isinstance(records, list) or not all(isinstance(rec, dict) for rec in records):
        logger.error(f"Invalid payload format for submission from {current_user_identity}: {records}")
        return jsonify({"msg": "Invalid data format. Expected an array of objects."}), 400

    try:
        data_to_insert = {
            "user_id": str(user['_id']), 
            "username": current_user_identity,
            "submission_timestamp": datetime.now(timezone.utc),
            "records": records,
            "status": "pending" 
        }
        
        # Create a copy for logging, converting datetime to string for JSON serialization
        loggable_data_to_insert = data_to_insert.copy()
        if isinstance(loggable_data_to_insert.get("submission_timestamp"), datetime):
            loggable_data_to_insert["submission_timestamp"] = loggable_data_to_insert["submission_timestamp"].isoformat()
        
        logger.info(f"Attempting to insert submission into submissions_collection for user {current_user_identity}: {json.dumps(loggable_data_to_insert, indent=2)}")
        
        result = submissions_collection.insert_one(data_to_insert) # Use original data_to_insert for MongoDB
        logger.info(f"Data submitted for approval successfully by {current_user_identity}. MongoDB Inserted ID: {result.inserted_id}")
        
        return jsonify({"msg": "Data submitted for approval successfully!", "requestId": str(data_to_insert['_id'])}), 201

    except Exception as e:
        logger.error(f"Error submitting data for user {current_user_identity} for approval: {str(e)}")
        return jsonify({"msg": f"Failed to submit data for approval: {str(e)}"}), 500


@app.route('/dashboard/download-excel', methods=['GET'])
@roles_required(['user', 'tester']) 
def download_daily_excel():
    """
    Downloads an Excel file of APPROVED data for the current day.
    Behavior depends on ENABLE_ALL_USERS_DAILY_DOWNLOAD config.
    """
    current_user_identity = get_jwt_identity() 
    user = users_collection.find_one({"username": current_user_identity})

    if not user:
        return jsonify({"msg": "User not found."}), 404

    try:
        today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        
        query_filter = {
            "submission_timestamp": {
                "$gte": today_start,
                "$lt": today_start + timedelta(days=1)
            },
            "status": "approved"
        }

        # Toggle logic for downloading
        if not ENABLE_ALL_USERS_DAILY_DOWNLOAD:
            query_filter["user_id"] = str(user['_id'])
            log_scope_msg = f"user {current_user_identity}'s specific"
            filename = f"{current_user_identity}_daily_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        else:
            log_scope_msg = "all users'"
            filename = f"all_users_daily_report_{datetime.now().strftime('%Y%m%d')}.xlsx"

        all_daily_entries = list(submissions_collection.find(query_filter))

        if not all_daily_entries:
            logger.info(f"No approved data found for today ({log_scope_msg} scope) for download requested by {current_user_identity}.")
            return jsonify({"msg": "No approved data found for today to download."}), 404

        records_to_download = []
        for entry in all_daily_entries:
            username_in_record = entry.get('username', 'N/A')
            submission_id = str(entry['_id'])
            submission_time_iso = entry.get('submission_timestamp', datetime.now(timezone.utc)).isoformat()

            if entry.get('records'):
                for record in entry['records']:
                    flattened_record = {
                        "Submitted By": username_in_record,
                        "Submission ID": submission_id,
                        "Submission Timestamp": submission_time_iso,
                        **record
                    }
                    records_to_download.append(flattened_record)

        if not records_to_download:
            logger.info(f"Found approved submissions but no valid records within them for today ({log_scope_msg} scope), requested by {current_user_identity}.")
            return jsonify({"msg": "No valid records found in approved submissions for today to download."}), 404

        all_keys = set()
        for record in records_to_download:
            all_keys.update(record.keys())
        
        normalized_records = []
        for record in records_to_download:
            normalized_record = {key: record.get(key) for key in all_keys}
            normalized_records.append(normalized_record)

        df = pd.DataFrame(normalized_records, columns=list(all_keys)) 

        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Daily Report', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Daily Report']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

        excel_buffer.seek(0)
        
        logger.info(f"Generated {log_scope_msg} daily Excel report for {current_user_identity}: {filename} with {len(records_to_download)} records.")

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Error generating {log_scope_msg} daily Excel report for {current_user_identity}: {str(e)}")
        return jsonify({"msg": f"Failed to generate daily report: {str(e)}"}), 500

@app.route('/dashboard/search-approved-data', methods=['GET'])
@roles_required(['user', 'tester']) 
def search_approved_data():
    """
    User endpoint to search APPROVED data and return generated barcodes for matching items.
    Behavior depends on ENABLE_ALL_USERS_GLOBAL_SEARCH config.
    """
    current_user_identity = get_jwt_identity()
    user = users_collection.find_one({"username": current_user_identity})

    if not user:
        return jsonify({"msg": "User not found."}), 404
    
    query_term = request.args.get('query', '').lower()
    
    if not query_term:
        return jsonify({"msg": "Search query is required."}), 400

    try:
        query_filter = {"status": "approved"}
        log_scope_msg = "all approved data (global)"
        
        if not ENABLE_ALL_USERS_GLOBAL_SEARCH:
            query_filter["user_id"] = str(user['_id'])
            log_scope_msg = f"user {current_user_identity}'s specific approved data"

        all_approved_submissions = list(submissions_collection.find(query_filter))
        
        matching_records = []
        for submission in all_approved_submissions:
            username_in_record = submission.get('username', 'N/A')
            submission_id = str(submission.get('_id'))
            submission_timestamp_iso = submission.get('submission_timestamp', datetime.now(timezone.utc)).isoformat()
            
            for record in submission.get('records', []):
                record_matches = False
                
                # Check for query_term in 'pid' or 'pname' specifically
                if (record.get('pid') and query_term in str(record['pid']).lower()) or \
                   (record.get('pname') and query_term in str(record['pname']).lower()):
                    record_matches = True
                
                # Also allow searching other string fields if needed (current behavior)
                if not record_matches: # Only check other fields if pid/pname didn't match
                    for key, value in record.items():
                        if isinstance(value, str) and query_term in value.lower():
                            record_matches = True
                            break 
                
                if record_matches:
                    upc_value = None
                    pid_str = str(record.get("pid", ""))
                    
                    # Attempt to derive an 11-digit numeric string from pid
                    numeric_pid = ''.join(filter(str.isdigit, pid_str))
                    
                    if len(numeric_pid) >= 11:
                        upc_value = numeric_pid[:11] # Take the first 11 digits
                    elif len(numeric_pid) > 0:
                        # Pad with leading zeros if less than 11 but has some digits
                        upc_value = numeric_pid.zfill(11)
                    else:
                        # Fallback for non-numeric or empty PIDs: Generate a stable 11-digit demo UPC
                        # Hash the original PID (or entire record as a fallback) to get a consistent number
                        # This is for DEMO PURPOSES ONLY, not for real unique UPC assignment
                        hash_input = pid_str if pid_str else json.dumps(record, separators=(',', ':'))
                        upc_hash = hashlib.sha256(hash_input.encode()).hexdigest()
                        upc_value = upc_hash[:11] # Take first 11 hex chars, then convert to digits
                        upc_value = ''.join(filter(str.isdigit, upc_value.zfill(11))) # Convert hex to digit, pad if needed
                        upc_value = upc_value[:11] # Ensure it's exactly 11 digits
                        logger.warning(f"PID '{pid_str}' not suitable for UPC. Generated demo UPC: {upc_value}")

                    if not upc_value: # Should not happen with current logic, but as a safeguard
                        upc_value = "01234567890" # Final fallback to a fixed demo UPC

                    barcode_svg_base64 = None
                    try:
                        # Generate UPCA barcode SVG
                        # UPCA expects an 11-digit string. It calculates the 12th checksum digit.
                        upca_barcode = UPCA(upc_value, writer=SVGWriter())
                        
                        svg_buffer = io.BytesIO()
                        # Adjust options for UPC-A visual style
                        upca_barcode.write(svg_buffer, options={
                            'module_height': 30, # Taller bars
                            'text_distance': 5, # Space between barcode and text
                            'font_size': 14,    # Font size for the numbers
                            'quiet_zone': 10    # Crucial white space around the barcode
                        }) 
                        svg_content = svg_buffer.getvalue().decode('utf-8')
                        
                        # Base64 encode the SVG content
                        barcode_svg_base64 = base64.b64encode(svg_content.encode('utf-8')).decode('utf-8')
                    except Exception as barcode_e:
                        logger.error(f"Error generating UPC-A barcode for record {record.get('pid', 'N/A')} with value '{upc_value}': {barcode_e}")
                        barcode_svg_base64 = None # If barcode generation fails, set it to None

                    flattened_record = {
                        "Submitted By": username_in_record,
                        "Submission ID": submission_id,
                        "Submission Timestamp": submission_timestamp_iso,
                        **record, # Include all original fields
                        "barcode_svg_base64": barcode_svg_base64, # Add the base64 barcode
                        "encoded_barcode_value": upca_barcode.get_fullcode() if barcode_svg_base64 else upc_value # Store the full 12-digit UPC including checksum
                    }
                    matching_records.append(flattened_record)
        
        if not matching_records:
            logger.info(f"User {current_user_identity} searched for '{query_term}', but no matching {log_scope_msg} data found.")
            return jsonify({"msg": "No matching approved data found."}), 404
        
        logger.info(f"User {current_user_identity} searched for '{query_term}'. Found {len(matching_records)} matching records in {log_scope_msg}.")
        return jsonify({"matching_records": matching_records}), 200

    except Exception as e:
        logger.error(f"Error searching {log_scope_msg} for user {current_user_identity}: {str(e)}")
        return jsonify({"msg": f"Failed to search approved data: {str(e)}"}), 500

@app.route('/dashboard/expiring-items', methods=['GET'])
@roles_required(['user', 'tester']) 
def get_expiring_items():
    """
    User: Get items that are expiring soon for the current user from approved submissions.
    Optional query parameter `days` specifies the notification threshold.
    This now includes items that have expired recently (within `days` threshold) and items expiring soon.
    """
    current_user_identity = get_jwt_identity()
    user = users_collection.find_one({"username": current_user_identity})

    if not user:
        return jsonify({"msg": "User not found."}), 404

    try:
        days_threshold = int(request.args.get('days', 30)) 
        if days_threshold < 0:
            return jsonify({'error': 'Days threshold cannot be negative.'}), 400
    except ValueError:
        return jsonify({'error': 'Invalid days threshold. Must be an integer.'}), 400

    # Define the date range for "expiring soon"
    # This now includes items that expired up to `days_threshold` days ago
    # and items that will expire up to `days_threshold` days from now.
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    past_limit = today_start - timedelta(days=days_threshold)
    future_limit = today_start + timedelta(days=days_threshold) # Inclusive of the last day

    logger.info(f"Checking expiring items for user {current_user_identity}: today={today_start.strftime('%d-%m-%Y')}, past_limit={past_limit.strftime('%d-%m-%Y')}, future_limit={future_limit.strftime('%d-%m-%Y')}")

    expiring_records = [] # Initialize here to collect all expiring records
    try: # <--- Start of the top-level try block for the function's main logic
        # Find approved submissions for the current user
        user_approved_submissions = submissions_collection.find({
            "user_id": str(user['_id']),
            "status": "approved"
        })
        
        found_user_submissions_flag = False # Use a flag to track if any submissions were found
        for submission in user_approved_submissions:
            found_user_submissions_flag = True # A submission was found
            logger.debug(f"Processing submission (ID: {str(submission.get('_id'))}) for user {current_user_identity}. Records count: {len(submission.get('records', []))}")

            for record in submission.get('records', []):
                expiry_date_str = record.get('expiry')
                logger.debug(f"  Checking record: {record.get('Item', 'N/A')}, raw expiry: '{expiry_date_str}'")

                if expiry_date_str:
                    try:
                        expiry_date = datetime.strptime(expiry_date_str, "%d-%m-%Y")
                        logger.debug(f"      Parsed expiry date: {expiry_date.strftime('%d-%m-%Y')}")
                        
                        if expiry_date >= past_limit and expiry_date <= future_limit:
                            expiring_records.append(record)
                            logger.info(f"      MATCH: Item '{record.get('Item', 'N/A')}: Expiry {expiry_date.strftime('%d-%m-%Y')} is within range.")
                        else:
                            logger.debug(f"      NO MATCH: Item '{record.get('Item', 'N/A')}: Expiry {expiry_date.strftime('%d-%m-%Y')} is OUTSIDE range.")
                    except ValueError:
                        logger.warning(f"      Could not parse expiry date '{expiry_date_str}' for record: {record.get('Item', 'N/A')} of user {current_user_identity}. Ensure 'expiry' is in dd-mm-yyyy format. Skipping record.")
                        continue
                else:
                    logger.debug(f"    Record {record.get('Item', 'N/A')} has no 'expiry' field. Skipping expiry check for this record.")
            
        if not found_user_submissions_flag: # If no submissions found for the user at all
            logger.info(f"No approved submissions found for user {current_user_identity}.")

        if expiring_records:
            return jsonify({'msg': f'{len(expiring_records)} items expiring soon.', 'expiring_items': expiring_records}), 200
        else:
            return jsonify({'msg': 'No items expiring soon found.'}), 200

    except Exception as e: # <--- End of the top-level try block
        logger.error(f"Error fetching expiring items for user {current_user_identity}: {str(e)}")
        return jsonify({'error': 'An unexpected error occurred while fetching expiring items.'}), 500


# --- Admin Endpoints (updated for approval workflow and single collection) ---
@app.route('/admin/pending-requests', methods=['GET'])
@admin_required 
def get_pending_requests():
    """
    Admin endpoint to fetch all pending data requests from submissions_collection.
    Requires admin role.
    """
    current_user_identity = get_jwt_identity()
    try:
        pending_requests = list(submissions_collection.find({"status": "pending"}))
        
        for req in pending_requests:
            req['_id'] = str(req['_id'])
            if 'submission_timestamp' in req and isinstance(req['submission_timestamp'], datetime):
                req['submission_timestamp'] = req['submission_timestamp'].isoformat()
        
        return jsonify({"pending_requests": pending_requests}), 200
    except Exception as e:
        logger.error(f"Admin user {current_user_identity} failed to fetch pending requests: {str(e)}")
        return jsonify({"msg": f"Failed to retrieve pending requests: {str(e)}"}), 500


@app.route('/admin/approve-request/<request_id>', methods=['POST'])
@admin_required 
def approve_request(request_id):
    """
    Admin endpoint to approve a pending data request.
    Updates submission status to "approved" in submissions_collection.
    """
    current_user_identity = get_jwt_identity()

    try:
        result = submissions_collection.update_one(
            {"_id": ObjectId(request_id), "status": "pending"},
            {"$set": {
                "status": "approved",
                "approved_by": current_user_identity,
                "approval_timestamp": datetime.now(timezone.utc)
            }}
        )

        if result.matched_count == 0:
            return jsonify({"msg": "Pending request not found or already processed."}), 404
        
        logger.info(f"Admin {current_user_identity} approved request {request_id}.")
        return jsonify({"msg": "Request approved!"}), 200

    except Exception as e:
        logger.error(f"Admin {current_user_identity} failed to approve request {request_id}: {str(e)}")
        return jsonify({"msg": f"Failed to approve request: {str(e)}"}), 500


@app.route('/admin/reject-request/<request_id>', methods=['POST'])
@admin_required 
def reject_request(request_id):
    """
    Admin endpoint to reject a pending data request.
    Updates status to "rejected" in submissions_collection.
    """
    current_user_identity = get_jwt_identity()

    try:
        result = submissions_collection.update_one(
            {"_id": ObjectId(request_id), "status": "pending"},
            {"$set": {
                "status": "rejected",
                "rejected_by": current_user_identity,
                "rejection_timestamp": datetime.now(timezone.utc)
            }}
        )

        if result.matched_count == 0:
            return jsonify({"msg": "Pending request not found or already processed."}), 404
        
        logger.info(f"Admin {current_user_identity} rejected request {request_id}.")
        return jsonify({"msg": "Request rejected."}), 200

    except Exception as e:
        logger.error(f"Admin {current_user_identity} failed to reject request {request_id}: {str(e)}")
        return jsonify({"msg": f"Failed to reject request: {str(e)}"}), 500


@app.route('/admin/daily-data', methods=['GET'])
@admin_required 
def get_all_daily_data():
    """
    Admin endpoint to fetch all APPROVED submissions from all users.
    Requires admin role.
    """
    current_user_identity = get_jwt_identity()

    try:
        # Fetch all submissions regardless of status
        all_daily_data = list(submissions_collection.find({})) # Changed to fetch all statuses
        
        for entry in all_daily_data:
            entry['_id'] = str(entry['_id'])
            if 'submission_timestamp' in entry and isinstance(entry['submission_timestamp'], datetime):
                entry['submission_timestamp'] = entry['submission_timestamp'].isoformat()
            if 'approval_timestamp' in entry and isinstance(entry['approval_timestamp'], datetime):
                entry['approval_timestamp'] = entry['approval_timestamp'].isoformat()
            if 'rejection_timestamp' in entry and isinstance(entry['rejection_timestamp'], datetime):
                entry['rejection_timestamp'] = entry['rejection_timestamp'].isoformat()
        
        return jsonify({"daily_data": all_daily_data}), 200
    except Exception as e:
        logger.error(f"Admin user {current_user_identity} failed to fetch all daily data: {str(e)}")
        return jsonify({"msg": f"Failed to retrieve all daily data: {str(e)}"}), 500

@app.route('/admin/download-all-excel', methods=['GET'])
@admin_required 
def download_all_daily_excel():
    """
    Admin endpoint to download all APPROVED submissions from all users as a single Excel file.
    Requires admin role.
    """
    current_user_identity = get_jwt_identity()

    try:
        all_approved_submissions = list(submissions_collection.find({"status": "approved"}))
        
        if not all_approved_submissions:
            return jsonify({"msg": "No approved data available to download across all users."}), 404

        flattened_records = []
        for entry in all_approved_submissions:
            # Using submission_timestamp for date_key if it exists, otherwise current time.
            # Using user_id directly from the submission document
            username = entry.get('username', 'N/A')
            timestamp = entry.get('submission_timestamp', datetime.now(timezone.utc)).isoformat()
            date_key = entry.get('submission_timestamp', datetime.now(timezone.utc)).isoformat().split('T')[0]
            submission_id = str(entry['_id']) # Use submission ID for clarity

            for record in entry.get('records', []):
                flattened_record = {
                    "Submission ID": submission_id,
                    "Date": date_key,
                    "Username": username,
                    "Submission Timestamp": timestamp,
                    **record 
                }
                flattened_records.append(flattened_record)
        
        if not flattened_records:
            return jsonify({"msg": "No valid records to convert to Excel."}), 404

        df = pd.DataFrame(flattened_records)
        excel_buffer = io.BytesIO()
        
        filename = f"all_approved_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='All Approved Data', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['All Approved Data']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

        excel_buffer.seek(0)
        
        logger.info(f"Admin user {current_user_identity} generated aggregated Excel report: {filename} with {len(flattened_records)} total records.")

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Admin user {current_user_identity} failed to generate aggregated Excel report: {str(e)}")
        return jsonify({"msg": f"Failed to generate aggregated Excel report: {str(e)}"}), 500

@app.route('/admin/clear-all-data', methods=['POST'])
@admin_required
def clear_all_data():
    """
    Admin endpoint to clear all data from the submissions_collection.
    Requires admin role.
    """
    current_user_identity = get_jwt_identity()

    try:
        result = submissions_collection.delete_many({})
        logger.info(f"Admin {current_user_identity} cleared {result.deleted_count} documents from submissions_collection.")
        return jsonify({"msg": f"Successfully cleared {result.deleted_count} data entries from all submissions."}), 200
    except Exception as e:
        logger.error(f"Admin {current_user_identity} failed to clear all submissions data: {str(e)}")
        return jsonify({"msg": f"Failed to clear all submissions data: {str(e)}"}), 500

@app.route('/admin/send-expired-notifications', methods=['POST'])
@admin_required
def send_expired_notifications():
    """
    Admin endpoint to send email notifications for expired/expiring items.
    Can target a specific user or all users.
    Payload: {"target_username": "user_to_notify" | "all", "days_threshold": int}
    This now includes items that have expired recently (within `days` threshold) and items expiring soon.
    """
    current_user_identity = get_jwt_identity()
    data = request.get_json()
    
    target_username = data.get('target_username', 'all')
    try:
        days_threshold = int(data.get('days_threshold', 30))
        if days_threshold < 0:
            return jsonify({'error': 'Days threshold cannot be negative.'}), 400
    except ValueError:
        return jsonify({'error': 'Invalid days threshold. Must be an integer.'}), 400

    # Define the date range for "expiring soon"
    # This now includes items that expired up to `days_threshold` days ago
    # and items that will expire up to `days_threshold` days from now.
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    past_limit = today_start - timedelta(days=days_threshold)
    future_limit = today_start + timedelta(days=days_threshold) # Inclusive of the last day

    logger.info(f"Sending expiry notifications (Admin: {current_user_identity}): target_username={target_username}, days_threshold={days_threshold}. Date range: [{past_limit.strftime('%d-%m-%Y')} - {future_limit.strftime('%d-%m-%Y')}]")
    
    users_to_notify = []
    if target_username == 'all':
        users_to_notify = list(users_collection.find({}))
        logger.info(f"Found {len(users_to_notify)} users for 'all' notification.")
    else:
        user = users_collection.find_one({"username": target_username})
        if user:
            users_to_notify.append(user)
            logger.info(f"Found specific user '{target_username}' for notification.")
        else:
            logger.warning(f"Target user '{target_username}' not found.")
            return jsonify({"msg": f"User '{target_username}' not found."}), 404

    successful_sends = []
    failed_sends = []

    try: # <--- Added top-level try block for the entire function
        for user in users_to_notify:
            user_email = user.get('email')
            logger.info(f"Processing user {user['username']} (Email: {user_email if user_email else 'N/A'})")

            if not user_email:
                logger.warning(f"Skipping email for user {user['username']}: No email address found.")
                failed_sends.append(f"User {user['username']} (no email)")
                continue

            expiring_records = []
            try:
                user_approved_submissions = submissions_collection.find({
                    "user_id": str(user['_id']),
                    "status": "approved"
                })
                
                found_user_submissions = False
                for submission in user_approved_submissions:
                    found_user_submissions = True # Corrected: Set to True if any submission is found
                    logger.debug(f"  Fetching records from submission ID: {str(submission.get('_id'))}")
                    for record in submission.get('records', []):
                        expiry_date_str = record.get('expiry')
                        logger.debug(f"    Record: {record.get('Item', 'N/A')}, raw expiry: '{expiry_date_str}'")

                        if expiry_date_str:
                            try:
                                expiry_date = datetime.strptime(expiry_date_str, "%d-%m-%Y")
                                logger.debug(f"      Parsed expiry date: {expiry_date.strftime('%d-%m-%Y')}")
                                
                                if expiry_date >= past_limit and expiry_date <= future_limit:
                                    expiring_records.append(record)
                                    logger.info(f"      MATCH: Item '{record.get('Item', 'N/A')}: Expiry {expiry_date.strftime('%d-%m-%Y')} is within range.")
                                else:
                                    logger.debug(f"      NO MATCH: Item '{record.get('Item', 'N/A')}: Expiry {expiry_date.strftime('%d-%m-%Y')} is OUTSIDE range.")
                            except ValueError:
                                logger.warning(f"      Could not parse expiry date '{expiry_date_str}' for record: {record.get('Item', 'N/A')} of user {current_user_identity}. Ensure 'expiry' is in dd-mm-yyyy format. Skipping record.")
                                continue
                        else:
                            logger.debug(f"    Record {record.get('Item', 'N/A')} has no 'expiry' field or it's empty. Skipping.")
                
                if not found_user_submissions: # If after looping through submissions, none were found for the user
                    logger.info(f"No approved submissions found for user {user['username']}.")

            except Exception as e:
                logger.error(f"Error fetching expiring items for user {user['username']}: {str(e)}")
                failed_sends.append(f"User {user['username']} (data fetch error)")
                continue

            if expiring_records:
                logger.info(f"For user {user['username']}, collected {len(expiring_records)} expiring records: {expiring_records}")
                subject = f"Urgent: Expiring/Expired Items Alert for {user['username']}"
                body_lines = [f"Dear {user['username']},\n",
                              f"The following items are expiring or have recently expired within the last {days_threshold} days or will expire in the next {days_threshold} days:\n"]
                for item in expiring_records:
                    # Dynamically construct item details for email body
                    item_details = ", ".join([f"{k}: {v}" for k, v in item.items()])
                    body_lines.append(f"- {item_details}")
                body_lines.append("\nPlease take action as necessary.")
                body_lines.append("\nBest regards,\nYour Excel Creator Team")
                
                email_body = "\n".join(body_lines)
                
                logger.info(f"Email body for {user['username']}:\n{email_body}")
                
                if send_email_notification(user_email, subject, email_body):
                    successful_sends.append(user['username'])
                else:
                    failed_sends.append(user['username'])
            else:
                logger.info(f"No expiring/recently expired items for user {user['username']}. Skipping email notification.")

        response_msg = { # This line should now be correctly handled by the outer try-except
            "msg": "Email notification process completed.",
            "successful_sends": successful_sends,
            "failed_sends": failed_sends
        }
        return jsonify(response_msg), 200

    except Exception as e: # <--- Catch-all for any other errors in the function
        logger.error(f"An unhandled error occurred in send_expired_notifications: {str(e)}")
        return jsonify({"msg": f"An unexpected error occurred: {str(e)}"}), 500


# --- NEW ADMIN ENDPOINT: Search User Submissions ---
@app.route('/admin/search-user-submissions', methods=['GET'])
@admin_required
def search_user_submissions():
    """
    Admin endpoint to fetch all submissions (pending, approved, rejected) for a specific user.
    Requires admin role.
    Query parameter: 'username'
    """
    current_user_identity = get_jwt_identity()
    target_username = request.args.get('username')

    if not target_username:
        logger.warning(f"Admin {current_user_identity} attempted user search without providing a username.")
        return jsonify({"msg": "Username query parameter is required."}), 400

    try:
        # Find submissions directly by username in the submissions_collection
        # This assumes the 'username' field is always present in submissions, which it should be.
        user_submissions_cursor = submissions_collection.find({"username": target_username})
        
        user_submissions = []
        for submission in user_submissions_cursor:
            submission['_id'] = str(submission['_id']) # Convert ObjectId to string
            # Convert datetime objects to ISO format strings for JSON serialization
            if 'submission_timestamp' in submission and isinstance(submission['submission_timestamp'], datetime):
                submission['submission_timestamp'] = submission['submission_timestamp'].isoformat()
            if 'approval_timestamp' in submission and isinstance(submission['approval_timestamp'], datetime):
                submission['approval_timestamp'] = submission['approval_timestamp'].isoformat()
            if 'rejection_timestamp' in submission and isinstance(submission['rejection_timestamp'], datetime):
                submission['rejection_timestamp'] = submission['rejection_timestamp'].isoformat()
            
            # Ensure records are treated as lists of dicts
            if 'records' in submission and not isinstance(submission['records'], list):
                submission['records'] = [] # Default to empty list if not valid
            
            user_submissions.append(submission)
        
        if not user_submissions:
            logger.info(f"Admin {current_user_identity}: No submissions found for user '{target_username}'.")
            return jsonify({"msg": f"No submissions found for user '{target_username}'."}), 404
        
        logger.info(f"Admin {current_user_identity}: Found {len(user_submissions)} submissions for user '{target_username}'.")
        return jsonify({"user_submissions": user_submissions}), 200

    except Exception as e:
        logger.error(f"Admin {current_user_identity} failed to search submissions for user '{target_username}': {str(e)}")
        return jsonify({"msg": f"Failed to retrieve data for user '{target_username}': {str(e)}"}), 500


if __name__ == '__main__':
    app.run(debug=True) # Run in debug mode for development
