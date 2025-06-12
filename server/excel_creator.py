from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_jwt_extended import create_access_token, jwt_required, JWTManager, get_jwt_identity, get_jwt
import pandas as pd
import io
import json
from datetime import datetime, timedelta, timezone
import logging

# For MongoDB
from pymongo import MongoClient
from bson.objectid import ObjectId # To handle MongoDB's default _id
from werkzeug.security import generate_password_hash, check_password_hash
import os # For environment variables

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# --- JWT Configuration ---
app.config["JWT_SECRET_KEY"] = os.getenv("JWT_SECRET_KEY", "super-secret-jwt-key") # Change this in production!
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = timedelta(hours=24) # Tokens valid for 24 hours
jwt = JWTManager(app)

# --- MongoDB Connection ---
# It's best practice to use environment variables for connection strings.
# Example: export MONGODB_URI="mongodb://localhost:27017/"
# from dotenv import load_dotenv
# load_dotenv() # Load environment variables from .env file

MONGO_URI = os.getenv('MONGODB_URI', 'mongodb://localhost:27017/') # For testing, replace with environment variable in production
client = MongoClient(MONGO_URI)
db = client.excel_creator_db # Your database name
templates_collection = db.templates # Collection for general templates
users_collection = db.users # Collection for users
daily_data_collection = db.daily_data # New collection for daily submitted data

# --- JWT Callbacks (for user identity) ---
@jwt.user_identity_loader
def user_identity_callback(user):
    return user['username']

@jwt.user_lookup_loader
def user_lookup_callback(_jwt_header, jwt_data):
    identity = jwt_data["sub"]
    return users_collection.find_one({"username": identity})

# --- Add some default templates and a default user to MongoDB if collections are empty ---
def initialize_db():
    if templates_collection.count_documents({}) == 0:
        logger.info("Initializing default templates into MongoDB...")
        default_templates = [
            {
                'id': 'sales_report',
                'name': 'Sales Report',
                'description': 'Basic sales data template',
                'columns': ['Product', 'Quantity', 'Price', 'Total'],
                'sample_data': [
                    {'Product': 'Laptop', 'Quantity': 5, 'Price': 999.99, 'Total': 4999.95},
                    {'Product': 'Mouse', 'Quantity': 20, 'Price': 25.50, 'Total': 510.00}
                ]
            },
            {
                'id': 'inventory',
                'name': 'Inventory List',
                'description': 'Product inventory tracking',
                'columns': ['Item', 'SKU', 'Stock', 'Location'],
                'sample_data': [
                    {'Item': 'Widget A', 'SKU': 'WA001', 'Stock': 150, 'Location': 'Warehouse A'},
                    {'Item': 'Widget B', 'SKU': 'WB001', 'Stock': 75, 'Location': 'Warehouse B'}
                ]
            },
            {
                'id': 'employees',
                'name': 'Employee List',
                'description': 'Basic employee information',
                'columns': ['Name', 'Department', 'Position', 'Email'],
                'sample_data': [
                    {'Name': 'John Doe', 'Department': 'IT', 'Position': 'Developer', 'Email': 'john@company.com'},
                    {'Name': 'Jane Smith', 'Department': 'HR', 'Position': 'Manager', 'Email': 'jane@company.com'}
                ]
            }
        ]
        templates_collection.insert_many(default_templates)
        logger.info("Default templates initialized.")
    else:
        logger.info("Templates collection is not empty, skipping initialization.")

    if users_collection.count_documents({}) == 0:
        logger.info("Initializing default users into MongoDB...")
        # Password for both is 'password'
        hashed_admin_password = generate_password_hash('admin')
        hashed_user_password = generate_password_hash('user')

        default_users = [
            {'username': 'admin', 'password': hashed_admin_password, 'role': 'admin'},
            {'username': 'user', 'password': hashed_user_password, 'role': 'user'}
        ]
        users_collection.insert_many(default_users)
        logger.info("Default users initialized.")
    else:
        logger.info("Users collection is not empty, skipping user initialization.")


# Call initialization on app startup
with app.app_context():
    initialize_db()


@app.route('/', methods=['GET'])
def home():
    """Health check endpoint"""
    return jsonify({
        'message': 'Excel Creator API is running',
        'status': 'healthy',
        'timestamp': datetime.now().isoformat()
    })

# --- Authentication Routes ---
@app.route('/login', methods=['POST'])
def login():
    username = request.json.get('username', None)
    password = request.json.get('password', None)

    if not username or not password:
        return jsonify({"msg": "Missing username or password"}), 400

    user = users_collection.find_one({"username": username})

    if user and check_password_hash(user['password'], password):
        access_token = create_access_token(identity=user)
        return jsonify(access_token=access_token, username=user['username'], role=user['role']), 200
    else:
        return jsonify({"msg": "Bad username or password"}), 401

@app.route('/register', methods=['POST'])
def register():
    username = request.json.get('username', None)
    password = request.json.get('password', None)
    role = request.json.get('role', 'user') # Default role is 'user'

    if not username or not password:
        return jsonify({"msg": "Missing username or password"}), 400

    if users_collection.find_one({"username": username}):
        return jsonify({"msg": "Username already exists"}), 409

    hashed_password = generate_password_hash(password)
    users_collection.insert_one({"username": username, "password": hashed_password, "role": role})

    return jsonify({"msg": "User registered successfully"}), 201

# --- Excel Creation & Template Routes (mostly unchanged, added JWT protection) ---
@app.route('/create-excel', methods=['POST'])
@jwt_required() # Protect this endpoint
def create_excel():
    """
    Create an Excel file from JSON data and return it as a downloadable file
    """
    try:
        current_user_identity = get_jwt_identity() # Get user identity from JWT
        logger.info(f"User {current_user_identity} requesting Excel creation.")

        json_data = request.get_json()
        
        if not json_data:
            return jsonify({'error': 'No JSON data provided'}), 400
        
        filename = json_data.get('filename', 'data')
        sheet_name = json_data.get('sheet_name', 'Sheet1')
        data = json_data.get('data', [])
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        if not isinstance(data, list):
            return jsonify({'error': 'Data must be a list of objects'}), 400
        
        for i, item in enumerate(data):
            if not isinstance(item, dict):
                return jsonify({'error': f'Item at index {i} is not a valid object'}), 400
        
        df = pd.DataFrame(data)
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
        
        excel_buffer.seek(0)
        safe_filename = f"{filename}.xlsx"
        
        logger.info(f"Created Excel file: {safe_filename} for user {current_user_identity} with {len(data)} rows")
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=safe_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error creating Excel file for user {current_user_identity}: {str(e)}")
        return jsonify({'error': f'Failed to create Excel file: {str(e)}'}), 500

@app.route('/validate-data', methods=['POST'])
@jwt_required() # Protect this endpoint
def validate_data():
    """
    Validate JSON data structure without creating the Excel file
    """
    try:
        json_data = request.get_json()
        
        if not json_data:
            return jsonify({'error': 'No JSON data provided'}), 400
        
        data = json_data.get('data', [])
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        if not isinstance(data, list):
            return jsonify({'error': 'Data must be a list of objects'}), 400
        
        columns = set()
        for i, item in enumerate(data):
            if not isinstance(item, dict):
                return jsonify({'error': f'Item at index {i} is not a valid object'}), 400
            columns.update(item.keys())
        
        df = pd.DataFrame(data)
        
        return jsonify({
            'valid': True,
            'row_count': len(data),
            'columns': list(columns),
            'preview': df.head().to_dict('records') if len(data) > 0 else []
        })
        
    except Exception as e:
        logger.error(f"Error validating data: {str(e)}")
        return jsonify({'error': f'Validation failed: {str(e)}'}), 500

@app.route('/templates', methods=['GET'])
@jwt_required() # Protect this endpoint
def get_templates():
    """
    Get available Excel templates from MongoDB
    """
    try:
        templates_cursor = templates_collection.find({}, {'_id': 0}) # Exclude _id
        templates = list(templates_cursor)
        return jsonify({'templates': templates})
    except Exception as e:
        logger.error(f"Error fetching templates from MongoDB: {str(e)}")
        return jsonify({'error': 'Failed to retrieve templates'}), 500

@app.route('/create-from-template/<template_id>', methods=['POST'])
@jwt_required() # Protect this endpoint
def create_from_template(template_id):
    """
    Create an Excel file from a predefined template stored in MongoDB
    """
    try:
        current_user_identity = get_jwt_identity()
        template = templates_collection.find_one({'id': template_id}, {'_id': 0})

        if not template:
            return jsonify({'error': 'Template not found'}), 404

        filename = template.get('name', 'template_data').lower().replace(' ', '_')
        sheet_name = template.get('sheet_name', template.get('name', 'Sheet1'))
        data = template.get('sample_data', []) # Use sample_data

        if not data:
            return jsonify({'error': 'Template has no sample data'}), 400

        df = pd.DataFrame(data)
        excel_buffer = io.BytesIO()

        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            from openpyxl.styles import Font, PatternFill, Border, Side
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

        excel_buffer.seek(0)
        safe_filename = f"{filename}.xlsx"

        logger.info(f"Created Excel file from template '{template_id}' for user {current_user_identity}: {safe_filename} with {len(data)} rows")

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=safe_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Error creating Excel file from template {template_id} for user {current_user_identity}: {str(e)}")
        return jsonify({'error': f'Failed to create Excel file from template: {str(e)}'}), 500

# --- NEW: Daily Data Submission & Download Endpoints ---

@app.route('/dashboard/submit-data', methods=['POST'])
@jwt_required()
def submit_daily_data():
    """
    Submits daily data for the authenticated user to MongoDB.
    The data is stored with a timestamp and linked to the user.
    Expected JSON payload: A list of objects (records).
    """
    current_user_identity = get_jwt_identity() # Get username from JWT
    user = users_collection.find_one({"username": current_user_identity})

    if not user:
        return jsonify({"msg": "User not found."}), 404
    
    records = request.get_json()

    if not records or not isinstance(records, list) or not all(isinstance(rec, dict) for rec in records):
        return jsonify({"msg": "Invalid data format. Expected an array of objects."}), 400

    try:
        # Get today's date in UTC and format it for consistent querying
        today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Prepare data for insertion
        data_to_insert = {
            "user_id": str(user['_id']), # Store MongoDB ObjectId as string
            "username": current_user_identity,
            "timestamp": datetime.now(timezone.utc),
            "date_key": today_start, # For easy daily lookup
            "records": records # The actual data submitted by the user
        }

        # Check if data already exists for this user and today
        existing_entry = daily_data_collection.find_one({
            "user_id": str(user['_id']),
            "date_key": today_start
        })

        if existing_entry:
            # If entry exists, append new records or update existing ones
            # For simplicity, let's append new records. A more complex logic
            # might involve merging or replacing based on specific keys.
            daily_data_collection.update_one(
                {"_id": existing_entry["_id"]},
                {"$push": {"records": {"$each": records}}} # Append new records
            )
            logger.info(f"Appended {len(records)} new records to daily data for user {current_user_identity} on {today_start.date()}.")
            return jsonify({"msg": "Data appended successfully!"}), 200
        else:
            # If no entry for today, insert a new one
            daily_data_collection.insert_one(data_to_insert)
            logger.info(f"Inserted new daily data for user {current_user_identity} on {today_start.date()} with {len(records)} records.")
            return jsonify({"msg": "Data submitted successfully!"}), 201

    except Exception as e:
        logger.error(f"Error submitting daily data for user {current_user_identity}: {str(e)}")
        return jsonify({"msg": f"Failed to submit data: {str(e)}"}), 500


@app.route('/dashboard/download-excel', methods=['GET'])
@jwt_required()
def download_daily_excel():
    """
    Downloads an Excel file of all data submitted by the authenticated user for the current day.
    """
    current_user_identity = get_jwt_identity() # Get username from JWT
    user = users_collection.find_one({"username": current_user_identity})

    if not user:
        return jsonify({"msg": "User not found."}), 404

    try:
        # Get today's date in UTC and format it for consistent querying
        today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Find daily data for the current user and today's date
        daily_entry = daily_data_collection.find_one({
            "user_id": str(user['_id']),
            "date_key": today_start
        })

        if not daily_entry or not daily_entry.get('records'):
            return jsonify({"msg": "No data found for today to download."}), 404

        records_to_download = daily_entry['records']

        # Ensure all records have a consistent set of columns for DataFrame creation
        all_keys = set()
        for record in records_to_download:
            all_keys.update(record.keys())
        
        # Create DataFrame, filling missing values for consistency
        df = pd.DataFrame(records_to_download, columns=list(all_keys))

        excel_buffer = io.BytesIO()
        filename = f"{current_user_identity}_daily_report_{datetime.now().strftime('%Y%m%d')}.xlsx"

        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Daily Report', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Daily Report']
            
            # Apply styling and auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

        excel_buffer.seek(0)
        
        logger.info(f"Generated daily Excel report for user {current_user_identity}: {filename} with {len(records_to_download)} records.")

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Error generating daily Excel report for user {current_user_identity}: {str(e)}")
        return jsonify({"msg": f"Failed to generate daily report: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True) # Run in debug mode for development
